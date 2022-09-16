import openpyxl
from datetime import datetime, date


def import_to_excel(json: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Факультеты')
    ws.append(('Id', 'Наименование'), )
    for row in json['faculties']:
        ws.append((row['Id'], row['Title']))

    ws = wb.create_sheet('Специализации')
    ws.append(('Id', 'Наименование', 'Id факультета', 'Наименование факультета'), )
    for row in json['specializations']:
        ws.append((row['Id'], row['Title'], row['Faculty']['Id'], row['Faculty']['Title']))

    ws = wb.create_sheet('Преподаватели')
    ws.append(('Id', 'Имя'), )
    for row in json['teachers']:
        ws.append((row['Id'], row['Name']))

    ws = wb.create_sheet('Курсы')
    ws.append(('Id', 'Наименование', 'Id факультета', 'Наименование факультета'), )
    for row in json['courses']:
        ws.append((row['Id'], row['Title'], row['Faculty']['Id'], row['Faculty']['Title']))

    ws = wb.create_sheet('Электронные почты')
    ws.append(('Id', 'Почта', 'Id студента'), )
    for row in json['emails']:
        ws.append((row['Id'], row['Mail'], row['StudentId']))

    ws = wb.create_sheet('Студенты')
    ws.append(('Id', 'Имя', 'Дата рождения', 'Id специализации', 'Наименование специализации',
               'Id факультета', 'Наименование факультета'), )
    for row in json['students']:
        ws.append((row['Id'], row['Name'], datetime.strptime(row['BirthDate'], '%Y-%m-%dT%H:%M:%S+05:00').date(),
                   row['Specialization']['Id'], row['Specialization']['Title'], row['Specialization']['Faculty']['Id'],
                   row['Specialization']['Faculty']['Title']))

    wb.save(f'{str(datetime.now()).replace(":", "-")}.xlsx')


if __name__ == '__main__':
    import_to_excel(
        {'faculties': [],
         'specializations': [],
         'teachers': [],
         'courses': [],
         'emails': [],
         'students': []})
